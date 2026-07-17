from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from src.calculations import calculate_batch_scenarios
from src.exports import (
    export_results_csv,
    export_results_excel,
    export_results_pdf,
    results_to_dataframe,
)
from src.models import FieldScenario


def _results():
    return calculate_batch_scenarios(
        [
            FieldScenario(field_id="a", field_name="A"),
            FieldScenario(field_id="b", field_name="B", drought_mode=True),
            FieldScenario(field_id="c", field_name="C", direct_soil_no3_n_ppm=50),
        ]
    )


def test_summary_dataframe_contains_inputs_intermediates_and_outputs() -> None:
    dataframe = results_to_dataframe(_results())
    required = {
        "field_id",
        "drought_mode",
        "direct_soil_no3_n_ppm",
        "soil_nitrate_input_unit",
        "direct_soil_nitrate_input_value",
        "soil_layers_json",
        "soil_no3_n_ppm",
        "organic_matter_pct",
        "total_n_credits_lb_ac",
        "standard_basal_n_need_lb_ac",
        "standard_unbounded_balance_lb_ac",
        "standard_fertilizer_recommendation_lb_ac",
        "full_yield_basal_n_need_lb_ac",
        "drought_adjusted_n_target_lb_ac",
        "drought_adjusted_fertilizer_recommendation_lb_ac",
    }
    assert required.issubset(dataframe.columns)
    assert len(dataframe) == 3


def test_csv_export_contains_all_records() -> None:
    text = export_results_csv(_results()).decode("utf-8-sig")
    assert "field_id" in text
    assert all(f",{name}," in text for name in ("A", "B", "C"))
    assert len(text.strip().splitlines()) == 4


def test_excel_export_has_expected_sheets_and_formatting() -> None:
    workbook = load_workbook(BytesIO(export_results_excel(_results())), data_only=False)
    assert workbook.sheetnames == ["Summary", "Inputs", "N Balance", "Methodology"]
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["Inputs"].max_row == 4
    assert workbook["N Balance"].max_row == 4
    assert workbook["N Balance"]["A1"].value == "field_id"
    assert workbook["N Balance"]["B2"].value == "A"
    assert workbook["Methodology"].max_row > 10
    assert workbook["Summary"]["A1"].font.bold


def test_pdf_export_contains_every_field_inputs_intermediates_and_references() -> None:
    pdf = export_results_pdf(_results())
    reader = PdfReader(BytesIO(pdf))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert pdf.startswith(b"%PDF-")
    assert len(reader.pages) >= 5
    assert "Batch summary" in text
    assert "Water-limited yield" in text
    assert "Drought fertilizer N" in text
    assert "Visual summaries" in text
    assert "Fertilizer recommendation comparison" in text
    assert "N source balance by field" in text
    assert all(name in text for name in ("A", "B", "C"))
    assert "Stable field ID: a" in text
    assert "N balance visual" in text
    assert "as context only" in text
    assert "User inputs" in text
    assert "Standard unbounded balance" in text
    assert "Experimental drought-adjusted N availability target" in text
    assert "Methodology and references" in text
    assert "Fertilizing Irrigated Corn" in text


def test_pdf_export_rejects_an_empty_batch() -> None:
    try:
        export_results_pdf([])
    except ValueError as exc:
        assert "At least one field result" in str(exc)
    else:
        raise AssertionError("Expected an empty PDF batch to be rejected")
