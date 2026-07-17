from __future__ import annotations

import json
import math
from datetime import date
from io import BytesIO
from xml.sax.saxutils import escape
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Flowable,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .calculations import calculate_soil_nitrate_lb_ac
from .models import FieldResult
from .references import METHODOLOGY_ROWS, REFERENCES
from .text import recommendation_summary


def _result_row(result: FieldResult) -> dict[str, object]:
    field = result.field
    drought = result.drought
    total_other = (
        result.credits.legume_n_lb_ac
        + result.credits.manure_n_lb_ac
        + result.credits.other_n_lb_ac
    )
    direct_soil_input_value = (
        field.direct_soil_no3_n_ppm
        if field.soil_nitrate_input_unit == "ppm"
        else calculate_soil_nitrate_lb_ac(field.direct_soil_no3_n_ppm, 24.0)
    )
    soil_layers_input = []
    for layer in field.soil_layers:
        thickness = layer.lower_depth_in - layer.upper_depth_in
        value = (
            layer.no3_n_ppm
            if field.soil_nitrate_input_unit == "ppm"
            else calculate_soil_nitrate_lb_ac(layer.no3_n_ppm, thickness)
        )
        soil_layers_input.append(
            {
                "upper_depth_in": layer.upper_depth_in,
                "lower_depth_in": layer.lower_depth_in,
                "nitrate_value": value,
                "unit": field.soil_nitrate_input_unit,
            }
        )
    return {
        "field_id": field.field_id,
        "field_name": field.field_name,
        "scenario_description": field.scenario_description,
        "calculation_mode": "standard + experimental drought" if drought else "standard CSU",
        "drought_mode": field.drought_mode,
        "full_yield_goal_bu_ac": field.expected_yield_bu_ac,
        "water_limited_yield_method": field.water_limited_yield_method,
        "yield_reduction_pct": field.yield_reduction_pct,
        "direct_water_limited_yield_bu_ac": field.direct_water_limited_yield_bu_ac,
        "water_limited_yield_bu_ac": drought.water_limited_yield_bu_ac if drought else None,
        "soil_nitrate_method": field.soil_nitrate_method,
        "soil_nitrate_input_unit": field.soil_nitrate_input_unit,
        "direct_soil_nitrate_input_value": direct_soil_input_value,
        "direct_soil_no3_n_ppm": field.direct_soil_no3_n_ppm,
        "soil_layers_input_json": json.dumps(soil_layers_input),
        "soil_layers_json": json.dumps([layer.__dict__ if hasattr(layer, "__dict__") else {"upper_depth_in": layer.upper_depth_in, "lower_depth_in": layer.lower_depth_in, "no3_n_ppm": layer.no3_n_ppm} for layer in field.soil_layers]),
        "soil_no3_n_ppm": result.soil_no3_n_ppm,
        "residual_n_credit_lb_ac": result.credits.residual_soil_n_lb_ac,
        "organic_matter_pct": field.organic_matter_pct,
        "organic_matter_credit_lb_ac": result.credits.organic_matter_n_lb_ac,
        "irrigation_no3_n_ppm": field.irrigation_no3_n_ppm,
        "irrigation_through_tasseling_ac_in": field.irrigation_through_tasseling_ac_in,
        "irrigation_n_credit_lb_ac": result.credits.irrigation_water_n_lb_ac,
        "legume_credit_lb_ac": result.credits.legume_n_lb_ac,
        "manure_credit_lb_ac": result.credits.manure_n_lb_ac,
        "other_n_credit_lb_ac": result.credits.other_n_lb_ac,
        "total_other_n_credits_lb_ac": total_other,
        "total_n_credits_lb_ac": result.credits.total_lb_ac,
        "standard_basal_n_need_lb_ac": result.standard.crop_n_need_lb_ac,
        "standard_unbounded_balance_lb_ac": result.standard.unbounded_balance_lb_ac,
        "standard_fertilizer_recommendation_lb_ac": result.standard.fertilizer_n_lb_ac,
        "donovan_reduction_pct": field.donovan_reduction_pct,
        "full_yield_basal_n_need_lb_ac": drought.full_yield_basal_n_need_lb_ac if drought else None,
        "drought_adjusted_n_target_lb_ac": drought.n_availability_target_lb_ac if drought else None,
        "drought_unbounded_balance_lb_ac": drought.unbounded_balance_lb_ac if drought else None,
        "drought_adjusted_fertilizer_recommendation_lb_ac": drought.fertilizer_n_lb_ac if drought else None,
        "validation_warnings": " | ".join(result.warnings),
    }


def results_to_dataframe(results: Iterable[FieldResult]) -> pd.DataFrame:
    return pd.DataFrame([_result_row(result) for result in results])


def export_results_csv(results: Iterable[FieldResult]) -> bytes:
    return results_to_dataframe(results).to_csv(index=False).encode("utf-8-sig")


def _write_dataframe(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1E4D2B")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        has_long_text = False
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if isinstance(value, float):
                cell.number_format = "0.00"
            if isinstance(value, str) and len(value) > 35:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                has_long_text = True
        if has_long_text:
            worksheet.row_dimensions[row_index].height = 60
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        values = [str(column_name)] + [str(value) for value in dataframe.iloc[:, column_index - 1].dropna().head(100)]
        width = min(45, max(12, max(map(len, values)) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.row_dimensions[1].height = 32


def export_results_excel(results: Iterable[FieldResult]) -> bytes:
    results_list = list(results)
    full = results_to_dataframe(results_list)
    summary_columns = [
        "field_id",
        "field_name",
        "scenario_description",
        "calculation_mode",
        "full_yield_goal_bu_ac",
        "water_limited_yield_bu_ac",
        "standard_basal_n_need_lb_ac",
        "drought_adjusted_n_target_lb_ac",
        "total_n_credits_lb_ac",
        "standard_fertilizer_recommendation_lb_ac",
        "drought_adjusted_fertilizer_recommendation_lb_ac",
    ]
    input_columns = [
        "field_id",
        "field_name",
        "scenario_description",
        "drought_mode",
        "full_yield_goal_bu_ac",
        "water_limited_yield_method",
        "yield_reduction_pct",
        "direct_water_limited_yield_bu_ac",
        "soil_nitrate_method",
        "soil_nitrate_input_unit",
        "direct_soil_nitrate_input_value",
        "direct_soil_no3_n_ppm",
        "soil_layers_input_json",
        "soil_layers_json",
        "soil_no3_n_ppm",
        "organic_matter_pct",
        "irrigation_no3_n_ppm",
        "irrigation_through_tasseling_ac_in",
        "legume_credit_lb_ac",
        "manure_credit_lb_ac",
        "other_n_credit_lb_ac",
        "donovan_reduction_pct",
    ]
    identifiers = ["field_id", "field_name", "scenario_description"]
    balance_columns = identifiers + [
        column for column in full.columns if column not in input_columns and column not in identifiers
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_dataframe(workbook, "Summary", full.reindex(columns=summary_columns))
    _write_dataframe(workbook, "Inputs", full.reindex(columns=input_columns))
    _write_dataframe(workbook, "N Balance", full.reindex(columns=balance_columns))

    methodology = workbook.create_sheet("Methodology")
    methodology.sheet_view.showGridLines = False
    methodology.append(["Calculation or assumption", "Equation / interpretation", "Primary source"])
    for row in METHODOLOGY_ROWS:
        methodology.append(row)
    methodology.append([])
    methodology.append(["Reference", "Local file", "External link"])
    for reference in REFERENCES:
        methodology.append([reference.title, reference.local_file, reference.url])
    methodology.append([])
    methodology.append(
        [
            "Limitation",
            "The experimental drought adjustment is a research-informed extrapolation and has not been validated as a CSU fertilizer recommendation algorithm.",
            "",
        ]
    )
    methodology.freeze_panes = "A2"
    methodology.column_dimensions["A"].width = 36
    methodology.column_dimensions["B"].width = 88
    methodology.column_dimensions["C"].width = 50
    for cell in methodology[1]:
        cell.fill = PatternFill("solid", fgColor="1E4D2B")
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in methodology[11]:
        cell.fill = PatternFill("solid", fgColor="1E4D2B")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in methodology.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, methodology.max_row + 1):
        longest = max(len(str(methodology.cell(row_index, column).value or "")) for column in range(1, 4))
        if longest > 100:
            methodology.row_dimensions[row_index].height = 48
        elif longest > 55:
            methodology.row_dimensions[row_index].height = 34

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_PDF_GREEN = colors.HexColor("#1E4D2B")
_PDF_GOLD = colors.HexColor("#C8C372")
_PDF_LIGHT_GREEN = colors.HexColor("#EFF6F0")
_PDF_LIGHT_GOLD = colors.HexColor("#FFF9E6")
_PDF_LIGHT_GRAY = colors.HexColor("#F4F5F4")
_PDF_TEXT = colors.HexColor("#26382C")
_PDF_SOURCE_COLORS = {
    "Residual soil nitrate": colors.HexColor("#4477AA"),
    "Soil organic matter": colors.HexColor("#228833"),
    "Irrigation water": colors.HexColor("#66CCEE"),
    "Previous crop / legume": colors.HexColor("#CCBB44"),
    "Manure": colors.HexColor("#AA3377"),
    "Other credit": colors.HexColor("#BBBBBB"),
    "Fertilizer required": colors.HexColor("#EE7733"),
}
_PDF_STANDARD_BLUE = colors.HexColor("#4477AA")
_PDF_DROUGHT_ORANGE = colors.HexColor("#EE7733")
_PDF_TARGET_PURPLE = colors.HexColor("#882255")


def _pdf_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "ReportTitle",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=_PDF_GREEN,
            alignment=TA_CENTER,
            spaceAfter=8,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=base["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#4B5563"),
            alignment=TA_CENTER,
            spaceAfter=14,
        ),
        "h1": ParagraphStyle(
            "ReportH1",
            parent=base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            textColor=_PDF_GREEN,
            spaceBefore=4,
            spaceAfter=8,
        ),
        "h2": ParagraphStyle(
            "ReportH2",
            parent=base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=_PDF_GREEN,
            spaceBefore=9,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "ReportBody",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=12,
            textColor=_PDF_TEXT,
            alignment=TA_LEFT,
        ),
        "small": ParagraphStyle(
            "ReportSmall",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=10,
            textColor=colors.HexColor("#4B5563"),
        ),
        "table_header": ParagraphStyle(
            "ReportTableHeader",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
        ),
        "metric": ParagraphStyle(
            "ReportMetric",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=_PDF_GREEN,
            alignment=TA_CENTER,
        ),
        "metric_label": ParagraphStyle(
            "ReportMetricLabel",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=7.5,
            leading=10,
            textColor=_PDF_TEXT,
            alignment=TA_CENTER,
        ),
    }


def _pdf_text(value: object) -> str:
    """Escape user-entered text and normalize generated punctuation for ReportLab."""
    text = str(value if value is not None else "")
    text = text.replace("\u2010", "-").replace("\u2011", "-")
    text = text.replace("\u2012", "-").replace("\u2013", "-").replace("\u2014", "-")
    text = text.replace("\u00d7", "x").replace("\u2022", "-")
    return escape(text).replace("\n", "<br/>")


def _pdf_paragraph(value: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(_pdf_text(value), style)


def _pdf_table(
    rows: list[list[object]],
    widths: list[float],
    styles: dict[str, ParagraphStyle],
    *,
    header: bool = True,
) -> Table:
    paragraph_rows = []
    for row_index, row in enumerate(rows):
        row_style = styles["table_header"] if header and row_index == 0 else styles["small"]
        paragraph_rows.append([_pdf_paragraph(value, row_style) for value in row])
    table = Table(paragraph_rows, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands: list[tuple] = [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CCD5CC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [colors.white, _PDF_LIGHT_GRAY]),
    ]
    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _PDF_GREEN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    table.setStyle(TableStyle(commands))
    return table


def _credit_pairs(result: FieldResult) -> list[tuple[str, float]]:
    c = result.credits
    return [
        ("Residual soil nitrate", c.residual_soil_n_lb_ac),
        ("Soil organic matter", c.organic_matter_n_lb_ac),
        ("Irrigation water", c.irrigation_water_n_lb_ac),
        ("Previous crop / legume", c.legume_n_lb_ac),
        ("Manure", c.manure_n_lb_ac),
        ("Other credit", c.other_n_lb_ac),
    ]


def _axis_max(value: float) -> float:
    if value <= 0:
        return 1.0
    target_step = value / 4.0
    for step in (5, 10, 20, 25, 50, 100, 200, 500):
        if step >= target_step:
            return math.ceil(value / step) * step
    magnitude = 10 ** math.floor(math.log10(value))
    step = 5 * magnitude
    return math.ceil(value / step) * step


def _chunks(items: list[FieldResult], size: int) -> Iterable[list[FieldResult]]:
    for index in range(0, len(items), size):
        yield items[index : index + size]


class _StackedHorizontalBarChart(Flowable):
    def __init__(
        self,
        rows: list[tuple[str, list[tuple[str, float]]]],
        *,
        markers: list[tuple[str, float, colors.Color]] | None = None,
        width: float = 6.5 * inch,
        bar_height: float = 14,
        row_gap: float = 8,
        label_width: float = 1.55 * inch,
    ) -> None:
        super().__init__()
        self.rows = rows
        self.markers = markers or []
        self.width = width
        self.bar_height = bar_height
        self.row_gap = row_gap
        self.label_width = label_width
        self.legend_items = [
            label
            for label in _PDF_SOURCE_COLORS
            if any(any(segment_label == label for segment_label, _ in segments) for _, segments in rows)
        ]
        self.legend_rows = max(1, math.ceil(len(self.legend_items) / 3))
        bars_height = len(rows) * bar_height + max(0, len(rows) - 1) * row_gap
        self.height = 24 + bars_height + 26 + self.legend_rows * 12

    def wrap(self, available_width: float, available_height: float) -> tuple[float, float]:
        self.width = min(self.width, available_width)
        return self.width, self.height

    def draw(self) -> None:
        canvas = self.canv
        plot_x = self.label_width
        plot_width = max(1, self.width - self.label_width - 0.18 * inch)
        legend_height = self.legend_rows * 12
        bars_bottom = legend_height + 26
        bars_top = self.height - 24
        max_value = max(
            [
                sum(value for _, value in segments)
                for _, segments in self.rows
            ]
            + [value for _, value, _ in self.markers]
            + [1.0]
        )
        axis_max = _axis_max(max_value * 1.08)

        canvas.setStrokeColor(colors.HexColor("#CCD5CC"))
        canvas.setLineWidth(0.5)
        canvas.line(plot_x, bars_bottom - 5, plot_x + plot_width, bars_bottom - 5)
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(colors.HexColor("#4B5563"))
        for tick in (0, axis_max / 2, axis_max):
            tick_x = plot_x + (tick / axis_max) * plot_width
            canvas.setStrokeColor(colors.HexColor("#D6DDD6"))
            canvas.line(tick_x, bars_bottom - 5, tick_x, bars_top)
            canvas.setFillColor(colors.HexColor("#4B5563"))
            canvas.drawCentredString(tick_x, bars_bottom - 17, f"{tick:.0f}")

        for marker_index, (label, value, color) in enumerate(self.markers):
            marker_x = plot_x + (value / axis_max) * plot_width
            canvas.setStrokeColor(color)
            canvas.setLineWidth(0.8)
            canvas.setDash(3, 2)
            canvas.line(marker_x, bars_bottom - 3, marker_x, bars_top + 2)
            canvas.setDash()
            canvas.setFillColor(color)
            canvas.setFont("Helvetica-Bold", 6.3)
            canvas.drawString(marker_x + 2, bars_top + 5 + 8 * marker_index, label)

        canvas.setFont("Helvetica", 7.2)
        for row_index, (row_label, segments) in enumerate(self.rows):
            y = bars_top - self.bar_height - row_index * (self.bar_height + self.row_gap)
            canvas.setFillColor(_PDF_TEXT)
            canvas.drawRightString(plot_x - 6, y + 4, row_label[:28])
            current_x = plot_x
            for segment_label, value in segments:
                if value <= 0:
                    continue
                segment_width = (value / axis_max) * plot_width
                canvas.setFillColor(_PDF_SOURCE_COLORS[segment_label])
                canvas.setStrokeColor(colors.white)
                canvas.rect(current_x, y, segment_width, self.bar_height, fill=1, stroke=1)
                if segment_width >= 22:
                    canvas.setFillColor(colors.white if segment_label != "Other credit" else _PDF_TEXT)
                    canvas.setFont("Helvetica-Bold", 6.2)
                    canvas.drawCentredString(current_x + segment_width / 2, y + 4, f"{value:.0f}")
                    canvas.setFont("Helvetica", 7.2)
                current_x += segment_width

        canvas.setFont("Helvetica", 6.5)
        for index, label in enumerate(self.legend_items):
            column = index % 3
            row = index // 3
            x = plot_x + column * (plot_width / 3)
            y = 4 + (self.legend_rows - 1 - row) * 12
            canvas.setFillColor(_PDF_SOURCE_COLORS[label])
            canvas.rect(x, y, 7, 7, fill=1, stroke=0)
            canvas.setFillColor(_PDF_TEXT)
            canvas.drawString(x + 10, y, label)


class _GroupedHorizontalBarChart(Flowable):
    def __init__(
        self,
        rows: list[tuple[str, float, float | None]],
        *,
        width: float = 6.5 * inch,
        label_width: float = 1.55 * inch,
    ) -> None:
        super().__init__()
        self.rows = rows
        self.width = width
        self.label_width = label_width
        self.row_height = 28
        self.height = 50 + len(rows) * self.row_height

    def wrap(self, available_width: float, available_height: float) -> tuple[float, float]:
        self.width = min(self.width, available_width)
        return self.width, self.height

    def draw(self) -> None:
        canvas = self.canv
        plot_x = self.label_width
        plot_width = max(1, self.width - self.label_width - 0.18 * inch)
        bars_bottom = 27
        bars_top = self.height - 22
        max_value = max(
            [
                value
                for _, standard, drought in self.rows
                for value in (standard, drought if drought is not None else 0)
            ]
            + [1.0]
        )
        axis_max = _axis_max(max_value * 1.12)

        canvas.setFont("Helvetica", 6.5)
        for tick in (0, axis_max / 2, axis_max):
            tick_x = plot_x + (tick / axis_max) * plot_width
            canvas.setStrokeColor(colors.HexColor("#D6DDD6"))
            canvas.setLineWidth(0.5)
            canvas.line(tick_x, bars_bottom, tick_x, bars_top)
            canvas.setFillColor(colors.HexColor("#4B5563"))
            canvas.drawCentredString(tick_x, 11, f"{tick:.0f}")

        for row_index, (field_name, standard, drought) in enumerate(self.rows):
            row_top = bars_top - row_index * self.row_height
            canvas.setFillColor(_PDF_TEXT)
            canvas.setFont("Helvetica", 7.2)
            canvas.drawRightString(plot_x - 6, row_top - 15, field_name[:28])
            for bar_index, (label, value, color) in enumerate(
                [
                    ("Standard", standard, _PDF_STANDARD_BLUE),
                    ("Drought", drought, _PDF_DROUGHT_ORANGE),
                ]
            ):
                if value is None:
                    continue
                y = row_top - 10 - bar_index * 11
                bar_width = (value / axis_max) * plot_width
                canvas.setFillColor(color)
                canvas.rect(plot_x, y, bar_width, 8, fill=1, stroke=0)
                canvas.setFillColor(_PDF_TEXT)
                canvas.setFont("Helvetica", 6.5)
                canvas.drawString(plot_x + bar_width + 3, y, f"{value:.1f}")

        legend_y = self.height - 12
        for index, (label, color) in enumerate(
            [("Standard CSU", _PDF_STANDARD_BLUE), ("Experimental drought", _PDF_DROUGHT_ORANGE)]
        ):
            x = plot_x + index * 1.7 * inch
            canvas.setFillColor(color)
            canvas.rect(x, legend_y, 7, 7, fill=1, stroke=0)
            canvas.setFillColor(_PDF_TEXT)
            canvas.setFont("Helvetica", 6.7)
            canvas.drawString(x + 10, legend_y, label)


def _individual_balance_chart(result: FieldResult) -> _StackedHorizontalBarChart:
    rows = [("Standard CSU", [*_credit_pairs(result), ("Fertilizer required", result.standard.fertilizer_n_lb_ac)])]
    markers = [("Standard crop N need", result.standard.crop_n_need_lb_ac, colors.HexColor("#222222"))]
    if result.drought:
        rows.append(
            (
                "Experimental drought",
                [*_credit_pairs(result), ("Fertilizer required", result.drought.fertilizer_n_lb_ac)],
            )
        )
        markers.append(("Experimental target", result.drought.n_availability_target_lb_ac, _PDF_TARGET_PURPLE))
    return _StackedHorizontalBarChart(rows, markers=markers, bar_height=16, row_gap=10)


def _batch_sources_chart(results: list[FieldResult]) -> _StackedHorizontalBarChart:
    rows = [
        (
            result.field.field_name or "Unnamed field",
            [*_credit_pairs(result), ("Fertilizer required", result.standard.fertilizer_n_lb_ac)],
        )
        for result in results
    ]
    return _StackedHorizontalBarChart(rows, bar_height=14, row_gap=8)


def _batch_recommendation_chart(results: list[FieldResult]) -> _GroupedHorizontalBarChart:
    return _GroupedHorizontalBarChart(
        [
            (
                result.field.field_name or "Unnamed field",
                result.standard.fertilizer_n_lb_ac,
                result.drought.fertilizer_n_lb_ac if result.drought else None,
            )
            for result in results
        ]
    )


def _format_number(value: float | None, suffix: str = "") -> str:
    return "-" if value is None else f"{value:.1f}{suffix}"


def _field_input_rows(result: FieldResult) -> list[list[object]]:
    field = result.field
    rows: list[list[object]] = [
        ["Input", "User-entered value"],
        ["Full-irrigation expected grain yield", _format_number(field.expected_yield_bu_ac, " bu/ac")],
        ["Experimental drought mode", "Enabled" if field.drought_mode else "Disabled"],
    ]
    if field.drought_mode:
        rows.append(
            [
                "Water-limited yield setup",
                "Percent reduction from full yield"
                if field.water_limited_yield_method == "percent_reduction"
                else "Direct water-limited yield",
            ]
        )
        if field.water_limited_yield_method == "percent_reduction":
            rows.append(["User-entered yield reduction", _format_number(field.yield_reduction_pct, "%")])
        else:
            rows.append(
                [
                    "User-entered water-limited yield",
                    _format_number(field.direct_water_limited_yield_bu_ac, " bu/ac"),
                ]
            )
        rows.append(
            [
                "Experimental reduction in optimum total N availability",
                _format_number(field.donovan_reduction_pct, "%"),
            ]
        )

    unit = "ppm NO3-N" if field.soil_nitrate_input_unit == "ppm" else "lb nitrate-N/ac"
    rows.extend(
        [
            [
                "Soil nitrate method",
                "Direct 0-24 inch weighted mean"
                if field.soil_nitrate_method == "direct"
                else "Sampled layers covering 0-24 inches",
            ],
            ["Soil nitrate input unit", unit],
        ]
    )
    if field.soil_nitrate_method == "direct":
        direct_value = (
            field.direct_soil_no3_n_ppm
            if field.soil_nitrate_input_unit == "ppm"
            else calculate_soil_nitrate_lb_ac(field.direct_soil_no3_n_ppm, 24.0)
        )
        rows.append(["Direct soil nitrate input", _format_number(direct_value, f" {unit}")])
    else:
        for index, layer in enumerate(field.soil_layers, start=1):
            thickness = layer.lower_depth_in - layer.upper_depth_in
            nitrate_value = (
                layer.no3_n_ppm
                if field.soil_nitrate_input_unit == "ppm"
                else calculate_soil_nitrate_lb_ac(layer.no3_n_ppm, thickness)
            )
            rows.append(
                [
                    f"Soil layer {index}: {layer.upper_depth_in:g}-{layer.lower_depth_in:g} inches",
                    _format_number(nitrate_value, f" {unit}"),
                ]
            )

    rows.extend(
        [
            ["Soil organic matter, 0-8 inches", _format_number(field.organic_matter_pct, "%")],
            ["Irrigation water NO3-N", _format_number(field.irrigation_no3_n_ppm, " ppm")],
            [
                "Irrigation applied through tasseling",
                _format_number(field.irrigation_through_tasseling_ac_in, " acre-inches"),
            ],
            ["Previous crop / legume N credit", _format_number(field.legume_credit_lb_ac, " lb N/ac")],
            ["Manure N credit", _format_number(field.manure_credit_lb_ac, " lb N/ac")],
            ["Other user-defined N credit", _format_number(field.other_n_credit_lb_ac, " lb N/ac")],
        ]
    )
    return rows


def _field_calculation_rows(result: FieldResult) -> list[list[object]]:
    rows: list[list[object]] = [
        ["Calculated component", "Value", "Role in calculation"],
        ["0-24 inch weighted mean soil NO3-N", _format_number(result.soil_no3_n_ppm, " ppm"), "Intermediate soil value"],
        ["Residual soil nitrate credit", _format_number(result.credits.residual_soil_n_lb_ac, " lb N/ac"), "N credit"],
        ["Soil organic matter credit", _format_number(result.credits.organic_matter_n_lb_ac, " lb N/ac"), "N credit"],
        ["Irrigation-water N credit", _format_number(result.credits.irrigation_water_n_lb_ac, " lb N/ac"), "N credit"],
        ["Previous crop / legume N credit", _format_number(result.credits.legume_n_lb_ac, " lb N/ac"), "N credit"],
        ["Manure N credit", _format_number(result.credits.manure_n_lb_ac, " lb N/ac"), "N credit"],
        ["Other user-defined N credit", _format_number(result.credits.other_n_lb_ac, " lb N/ac"), "N credit"],
        ["Total N credits", _format_number(result.credits.total_lb_ac, " lb N/ac"), "Estimated N availability credited"],
        ["Standard basal crop N need", _format_number(result.standard.crop_n_need_lb_ac, " lb N/ac"), "Crop N need; not fertilizer or uptake"],
        ["Standard unbounded balance", _format_number(result.standard.unbounded_balance_lb_ac, " lb N/ac"), "Crop N need minus all credits"],
        ["Standard CSU fertilizer N recommendation", _format_number(result.standard.fertilizer_n_lb_ac, " lb N/ac"), "Unbounded balance, bounded at zero"],
    ]
    if result.drought:
        rows.extend(
            [
                ["Water-limited yield goal", _format_number(result.drought.water_limited_yield_bu_ac, " bu/ac"), "Scenario yield"],
                ["Full-yield basal crop N need", _format_number(result.drought.full_yield_basal_n_need_lb_ac, " lb N/ac"), "Full-water target basis for Donovan adjustment"],
                ["Experimental drought-adjusted N availability target", _format_number(result.drought.n_availability_target_lb_ac, " lb N/ac"), "Experimental target; not plant uptake"],
                ["Experimental unbounded balance", _format_number(result.drought.unbounded_balance_lb_ac, " lb N/ac"), "Experimental target minus all credits"],
                ["Experimental fertilizer N recommendation", _format_number(result.drought.fertilizer_n_lb_ac, " lb N/ac"), "Unbounded balance, bounded at zero"],
            ]
        )
    return rows


def _draw_pdf_page(canvas, document) -> None:
    canvas.saveState()
    canvas.setStrokeColor(_PDF_GOLD)
    canvas.setLineWidth(1.5)
    canvas.line(document.leftMargin, 0.56 * inch, letter[0] - document.rightMargin, 0.56 * inch)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor("#4B5563"))
    canvas.drawString(document.leftMargin, 0.37 * inch, "Colorado Corn Nitrogen Planner")
    canvas.drawRightString(letter[0] - document.rightMargin, 0.37 * inch, f"Page {document.page}")
    canvas.restoreState()


def export_results_pdf(results: Iterable[FieldResult]) -> bytes:
    """Create one printable report containing every supplied field result."""
    results_list = list(results)
    if not results_list:
        raise ValueError("At least one field result is required for a PDF report.")

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.6 * inch,
        leftMargin=0.6 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.75 * inch,
        title="Colorado Corn Nitrogen Planner - Batch Report",
        author="Colorado Corn Nitrogen Planner",
        subject="Field inputs, calculations, and nitrogen recommendations",
    )
    styles = _pdf_styles()
    story: list[object] = [
        _pdf_paragraph("Colorado Corn Nitrogen Planner", styles["title"]),
        _pdf_paragraph("Batch recommendation report", styles["h1"]),
        _pdf_paragraph(
            f"Generated {date.today().isoformat()} - {len(results_list)} field scenario(s)",
            styles["subtitle"],
        ),
        _pdf_paragraph(
            "This report records the user inputs, intermediate calculations, and final recommendations "
            "for every field scenario included in the batch. Crop N need, estimated N availability, "
            "fertilizer N, and plant N uptake are distinct quantities.",
            styles["body"],
        ),
        Spacer(1, 0.14 * inch),
        _pdf_paragraph("Batch summary", styles["h1"]),
    ]

    summary_rows: list[list[object]] = [
        ["Field", "Full yield", "Water-limited yield", "Standard fertilizer N", "Drought fertilizer N"]
    ]
    for result in results_list:
        summary_rows.append(
            [
                result.field.field_name or "Unnamed field",
                _format_number(result.field.expected_yield_bu_ac, " bu/ac"),
                _format_number(result.drought.water_limited_yield_bu_ac, " bu/ac") if result.drought else "-",
                _format_number(result.standard.fertilizer_n_lb_ac, " lb/ac"),
                _format_number(result.drought.fertilizer_n_lb_ac, " lb/ac") if result.drought else "-",
            ]
        )
    story.append(
        _pdf_table(
            summary_rows,
            [1.55 * inch, 1.05 * inch, 1.25 * inch, 1.35 * inch, 1.3 * inch],
            styles,
        )
    )
    story.extend(
        [
            Spacer(1, 0.14 * inch),
            _pdf_paragraph(
                "Decision-support and education only. Confirm recommendations with current soil and water "
                "tests and current local agronomic advice.",
                styles["small"],
            ),
            Spacer(1, 0.14 * inch),
            _pdf_paragraph("Visual summaries", styles["h1"]),
        ]
    )
    for chunk_index, chunk in enumerate(_chunks(results_list, 8), start=1):
        chunk_label = f" - fields {1 + (chunk_index - 1) * 8}-{(chunk_index - 1) * 8 + len(chunk)}" if len(results_list) > 8 else ""
        story.extend(
            [
                _pdf_paragraph(f"Fertilizer recommendation comparison{chunk_label}", styles["h2"]),
                _batch_recommendation_chart(chunk),
                Spacer(1, 0.12 * inch),
                _pdf_paragraph(f"N source balance by field{chunk_label}", styles["h2"]),
                _batch_sources_chart(chunk),
                Spacer(1, 0.12 * inch),
            ]
        )

    for result in results_list:
        field = result.field
        story.extend(
            [
                PageBreak(),
                _pdf_paragraph(field.field_name or "Unnamed field", styles["h1"]),
                _pdf_paragraph(f"Stable field ID: {field.field_id}", styles["small"]),
            ]
        )
        if field.scenario_description:
            story.extend(
                [
                    Spacer(1, 0.06 * inch),
                    _pdf_paragraph(f"Scenario description: {field.scenario_description}", styles["body"]),
                ]
            )

        standard_metric = Table(
            [
                [_pdf_paragraph("STANDARD CSU FERTILIZER N", styles["metric_label"])],
                [_pdf_paragraph(f"{result.standard.fertilizer_n_lb_ac:.1f} lb N/ac", styles["metric"])],
            ],
            colWidths=[3.05 * inch],
        )
        standard_metric.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), _PDF_LIGHT_GREEN),
                    ("BOX", (0, 0), (-1, -1), 1.5, _PDF_GREEN),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        if result.drought:
            drought_metric = Table(
                [
                    [_pdf_paragraph("EXPERIMENTAL DROUGHT FERTILIZER N", styles["metric_label"])],
                    [_pdf_paragraph(f"{result.drought.fertilizer_n_lb_ac:.1f} lb N/ac", styles["metric"])],
                ],
                colWidths=[3.05 * inch],
            )
            drought_metric.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), _PDF_LIGHT_GOLD),
                        ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#8A6D1D")),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            metric_table = Table([[standard_metric, drought_metric]], colWidths=[3.25 * inch, 3.25 * inch])
        else:
            metric_table = Table([[standard_metric]], colWidths=[3.25 * inch], hAlign="LEFT")
        metric_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.extend(
            [
                Spacer(1, 0.12 * inch),
                metric_table,
                _pdf_paragraph("N balance visual", styles["h2"]),
                _individual_balance_chart(result),
                _pdf_paragraph("Interpretation", styles["h2"]),
                _pdf_paragraph(recommendation_summary(result), styles["body"]),
                _pdf_paragraph("User inputs", styles["h2"]),
                _pdf_table(_field_input_rows(result), [2.75 * inch, 3.75 * inch], styles),
                KeepTogether(
                    [
                        _pdf_paragraph("Intermediate calculations and recommendations", styles["h2"]),
                        _pdf_table(
                            _field_calculation_rows(result),
                            [2.65 * inch, 1.35 * inch, 2.5 * inch],
                            styles,
                        ),
                    ]
                ),
            ]
        )
        if result.warnings:
            story.extend(
                [
                    _pdf_paragraph("Validation warnings", styles["h2"]),
                    _pdf_paragraph(" | ".join(result.warnings), styles["body"]),
                ]
            )

    story.extend(
        [
            PageBreak(),
            _pdf_paragraph("Methodology and references", styles["h1"]),
            _pdf_paragraph(
                "Equations and assumptions are listed here so each reported value remains traceable to "
                "the documented methodology.",
                styles["body"],
            ),
            Spacer(1, 0.08 * inch),
            _pdf_table(
                [["Calculation or assumption", "Equation / interpretation", "Primary source"], *[list(row) for row in METHODOLOGY_ROWS]],
                [1.65 * inch, 3.45 * inch, 1.4 * inch],
                styles,
            ),
            _pdf_paragraph("Scientific references", styles["h2"]),
        ]
    )
    for reference in REFERENCES:
        story.append(
            Paragraph(
                f"<b>{_pdf_text(reference.title)}</b><br/>{_pdf_text(reference.authors)}<br/>"
                f"Role: {_pdf_text(reference.role)}<br/>Local source: {_pdf_text(reference.local_file)}<br/>"
                f"External source: <link href=\"{escape(reference.url)}\" color=\"#1E4D2B\">{_pdf_text(reference.url)}</link>",
                styles["small"],
            )
        )
        story.append(Spacer(1, 0.08 * inch))
    story.extend(
        [
            _pdf_paragraph("Important limitation", styles["h2"]),
            _pdf_paragraph(
                "The experimental drought adjustment is a research-informed extrapolation and has not "
                "been validated as a CSU fertilizer recommendation algorithm. Standard CSU results remain "
                "separate and visible for every drought scenario.",
                styles["body"],
            ),
        ]
    )

    document.build(story, onFirstPage=_draw_pdf_page, onLaterPages=_draw_pdf_page)
    return buffer.getvalue()
